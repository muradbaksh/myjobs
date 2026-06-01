import csv
from openpyxl import Workbook
from django.http import HttpResponse

from reviews.models import CompanyReview
from compensation.models import Compensation


class ExportService:

    @staticmethod
    def export_reviews_csv():

        response = HttpResponse(
            content_type="text/csv"
        )

        response["Content-Disposition"] = (
            'attachment; filename="reviews.csv"'
        )

        writer = csv.writer(response)

        writer.writerow([
            "Company",
            "Score",
            "Recommendation",
            "Created At",
        ])

        reviews = CompanyReview.objects.select_related(
            "company"
        )

        for review in reviews:

            writer.writerow([
                review.company.name,
                review.weighted_score,
                review.overall_recommendation,
                review.created_at,
            ])

        return response

    @staticmethod
    def export_compensation_excel():

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Compensation"

        headers = [
            "Company",
            "Job Title",
            "Experience",
            "Location",
            "Total Compensation",
        ]

        for col_num, header in enumerate(
            headers,
            start=1
        ):
            sheet.cell(
                row=1,
                column=col_num,
                value=header
            )

        compensations = (
            Compensation.objects.select_related(
                "company"
            )
        )

        row = 2

        for item in compensations:

            sheet.cell(
                row=row,
                column=1,
                value=item.company.name
            )

            sheet.cell(
                row=row,
                column=2,
                value=item.normalized_job_title
            )

            sheet.cell(
                row=row,
                column=3,
                value=item.experience_level
            )

            sheet.cell(
                row=row,
                column=4,
                value=item.location
            )

            sheet.cell(
                row=row,
                column=5,
                value=float(item.total_compensation)
            )

            row += 1

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )

        response["Content-Disposition"] = (
            'attachment; filename="compensation.xlsx"'
        )

        workbook.save(response)

        return response